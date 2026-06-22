"""
ETL / Conversion Engine
Joins platform order data against the master mapping, converts ordered
units into Amul SAP box/case quantities (rounding up on any remainder
and flagging it), and builds the manager-facing monthly projection
output with one Qty + Date column pair per platform.
"""

import math
import pandas as pd
from mapping_engine import load_master_mapping


def convert_platform_orders(platform_orders: pd.DataFrame, platform: str) -> dict:
    """
    Maps one platform's normalized order rows to SAP codes and computes
    box quantities.
    Returns dict with: mapped (df), unmapped (df), stats.
    """
    mapping = load_master_mapping()
    plat_map = mapping[mapping["platform"] == platform][
        ["platform_sku", "sap_code", "sap_description", "fg_group",
         "case_pack_size", "pack_confidence"]
    ]

    merged = platform_orders.merge(plat_map, on="platform_sku", how="left")

    is_mapped = merged["sap_code"].notna()
    mapped = merged[is_mapped].copy()
    unmapped = merged[~is_mapped].copy()

    if len(mapped):
        mapped["case_pack_size"] = mapped["case_pack_size"].fillna(1)
        raw_boxes = mapped["order_qty_units"] / mapped["case_pack_size"]
        mapped["po_qty_boxes"] = raw_boxes.apply(math.ceil)
        mapped["had_remainder"] = (mapped["order_qty_units"] % mapped["case_pack_size"]) != 0
        mapped["remainder_units"] = mapped["order_qty_units"] - (
            (mapped["po_qty_boxes"] - mapped["had_remainder"].astype(int)) * mapped["case_pack_size"]
        )
        mapped["platform"] = platform

    stats = {
        "platform": platform,
        "total_rows": len(merged),
        "mapped_rows": len(mapped),
        "unmapped_rows": len(unmapped),
        "rounded_up_rows": int(mapped["had_remainder"].sum()) if len(mapped) else 0,
        "total_units": float(platform_orders["order_qty_units"].sum()),
        "total_boxes": float(mapped["po_qty_boxes"].sum()) if len(mapped) else 0.0,
    }
    return {"mapped": mapped, "unmapped": unmapped, "stats": stats}


def build_manager_projection(all_mapped: dict) -> pd.DataFrame:
    """
    Builds the manager-facing wide table:
    SAP Code | SAP Description | FG Group |
    [Platform] PO Qty (Boxes) | [Platform] Order Date | [Platform] Rounded Up?
    ... repeated per platform ...
    | Total PO Qty (Boxes)

    all_mapped: dict of {platform_name: mapped_dataframe}
    """
    base_cols = ["sap_code", "sap_description", "fg_group"]
    platform_frames = []

    for platform, df in all_mapped.items():
        if df.empty:
            continue
        grp = (
            df.groupby(base_cols, as_index=False)
            .agg(
                po_qty_boxes=("po_qty_boxes", "sum"),
                order_date=("order_date", "max"),
                had_remainder=("had_remainder", "any"),
            )
        )
        grp = grp.rename(columns={
            "po_qty_boxes": f"{platform} PO Qty (Boxes)",
            "order_date": f"{platform} Order Date",
            "had_remainder": f"{platform} Rounded Up",
        })
        platform_frames.append(grp)

    if not platform_frames:
        return pd.DataFrame()

    result = platform_frames[0]
    for pf in platform_frames[1:]:
        result = result.merge(pf, on=base_cols, how="outer")

    qty_cols = [c for c in result.columns if "PO Qty (Boxes)" in c]
    result["Total PO Qty (Boxes)"] = result[qty_cols].fillna(0).sum(axis=1)

    result = result.rename(columns={
        "sap_code": "SAP Code",
        "sap_description": "SAP Product Description",
        "fg_group": "FG Group",
    })

    # Order columns: base, then per-platform (Qty, Date, RoundedUp) groups, then total
    ordered_cols = ["SAP Code", "SAP Product Description", "FG Group"]
    for platform in all_mapped.keys():
        for suffix in ["PO Qty (Boxes)", "Order Date", "Rounded Up"]:
            col = f"{platform} {suffix}"
            if col in result.columns:
                ordered_cols.append(col)
    ordered_cols.append("Total PO Qty (Boxes)")
    ordered_cols = [c for c in ordered_cols if c in result.columns]

    return result[ordered_cols].sort_values("SAP Code").reset_index(drop=True)


def export_projection_to_excel(projection_df: pd.DataFrame, unmapped_by_platform: dict,
                                stats_by_platform: list) -> bytes:
    """Builds the final downloadable Excel with Summary, Projection (rounded rows highlighted),
    and Unmapped SKUs sheets."""
    import io
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(stats_by_platform).to_excel(writer, sheet_name="Summary", index=False)
        projection_df.to_excel(writer, sheet_name="Monthly PO Projection", index=False)

        unmapped_frames = []
        for platform, df in unmapped_by_platform.items():
            if not df.empty:
                tmp = df[["platform_sku", "raw_product_name", "order_qty_units"]].copy()
                tmp.insert(0, "platform", platform)
                unmapped_frames.append(tmp)
        if unmapped_frames:
            pd.concat(unmapped_frames, ignore_index=True).to_excel(
                writer, sheet_name="Unmapped SKUs", index=False
            )

        # Highlight "Rounded Up" = True cells and their matching qty cell
        ws = writer.sheets["Monthly PO Projection"]
        yellow = PatternFill("solid", start_color="FFF6CC", end_color="FFF6CC")
        bold_orange = Font(color="B45309", bold=True)
        header = [c.value for c in ws[1]]

        rounded_cols = [i for i, h in enumerate(header, start=1) if h and "Rounded Up" in h]
        for col_idx in rounded_cols:
            qty_col_idx = col_idx - 2  # Qty column sits two columns before "Rounded Up" in our ordering
            col_letter_flag = get_column_letter(col_idx)
            col_letter_qty = get_column_letter(qty_col_idx)
            for row in range(2, ws.max_row + 1):
                val = ws[f"{col_letter_flag}{row}"].value
                if val is True:
                    ws[f"{col_letter_qty}{row}"].fill = yellow
                    ws[f"{col_letter_qty}{row}"].font = bold_orange

        # Auto column width across all sheets
        for sheet in writer.sheets.values():
            for col in sheet.columns:
                max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
                sheet.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

    return output.getvalue()
